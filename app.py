import os
import io
import uuid
import smtplib
import qrcode
from datetime import datetime, date
from functools import wraps
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, session, jsonify, send_file, send_from_directory
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PIL import Image as PILImage, ImageDraw, ImageFont

app = Flask(__name__)
app.secret_key = "absensi_event_secret_2026"
app.config['SESSION_PERMANENT'] = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
QR_DIR = os.path.join(BASE_DIR, "qrcodes")
PESERTA_FILE = os.path.join(DATA_DIR, "peserta.xlsx")
CONFIG_FILE = os.path.join(DATA_DIR, "config.xlsx")
SESI_FILE = os.path.join(DATA_DIR, "sesi.xlsx")
ABSENSI_FILE = os.path.join(DATA_DIR, "absensi.xlsx")
AUDIT_LOG_FILE = os.path.join(DATA_DIR, "audit_log.xlsx")

ADMIN_USER = "bangJago"
ADMIN_PASS = "mudahditebak"

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(QR_DIR, exist_ok=True)


def init_excel():
    if not os.path.exists(PESERTA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Peserta"
        ws.append(["ID", "Nama", "Email", "NoHP", "Divisi", "QRCode", "Hadir", "WaktuCheckIn", "Tanggal"])
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        wb.save(PESERTA_FILE)

    if not os.path.exists(CONFIG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Key", "Value"])
        settings = [
            ["event_name", "Nama Event"],
            ["checkin_start", "08:00"],
            ["checkin_end", "10:00"],
            ["checkout_start", "16:00"],
            ["checkout_end", "18:00"],
            ["notif_sound", "true"],
        ]
        for s in settings:
            ws.append(s)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        wb.save(CONFIG_FILE)

    if not os.path.exists(SESI_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sesi"
        ws.append(["ID", "Nama", "Aktif"])
        ws.append([1, "Absen Awal", True])
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        wb.save(SESI_FILE)

    if not os.path.exists(ABSENSI_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Absensi"
        ws.append(["IDPeserta", "NamaPeserta", "Divisi", "IDSesi", "Waktu", "Tanggal"])
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        wb.save(ABSENSI_FILE)

    if not os.path.exists(AUDIT_LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "AuditLog"
        ws.append(["Waktu", "IDPeserta", "NamaPeserta", "Aksi", "Sesi", "Status", "Detail"])
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        wb.save(AUDIT_LOG_FILE)


init_excel()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


def get_config():
    config = {}
    if os.path.exists(CONFIG_FILE):
        wb = load_workbook(CONFIG_FILE)
        ws = wb["Settings"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                config[row[0]] = row[1] or ""
        wb.close()
    return config


def save_config(config):
    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.append(["Key", "Value"])
    for key, value in config.items():
        ws.append([key, value])
    wb.save(CONFIG_FILE)


def get_peserta_list():
    peserta = []
    if os.path.exists(PESERTA_FILE):
        wb = load_workbook(PESERTA_FILE)
        ws = wb["Peserta"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                peserta.append({
                    "id": row[0],
                    "nama": row[1],
                    "email": row[2],
                    "nohp": row[3],
                    "divisi": row[4],
                    "qrcode": row[5],
                    "hadir": row[6],
                    "waktu_checkin": row[7],
                    "tanggal": row[8],
                })
        wb.close()
    return peserta


def get_next_id():
    peserta = get_peserta_list()
    if not peserta:
        return "PST-001"
    nums = []
    for p in peserta:
        try:
            num = int(p["id"].split("-")[1])
            nums.append(num)
        except (IndexError, ValueError):
            pass
    if nums:
        return f"PST-{max(nums) + 1:03d}"
    return "PST-001"


def add_peserta(nama, email, nohp, divisi):
    pid = get_next_id()
    qr_code = str(uuid.uuid4()).replace("-", "")[:12].upper()

    wb = load_workbook(PESERTA_FILE)
    ws = wb["Peserta"]
    ws.append([pid, nama, email, nohp, divisi, qr_code, False, "", ""])
    wb.save(PESERTA_FILE)

    generate_qr_image(pid, qr_code, nama)
    return pid, qr_code


def delete_peserta(pid):
    wb = load_workbook(PESERTA_FILE)
    ws = wb["Peserta"]
    rows_to_delete = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == pid:
            rows_to_delete.append(idx)
            qr_path = os.path.join(QR_DIR, f"{row[5]}.png")
            if os.path.exists(qr_path):
                os.remove(qr_path)
    for idx in reversed(rows_to_delete):
        ws.delete_rows(idx)
    wb.save(PESERTA_FILE)


def generate_qr_image(pid, qr_code, nama):
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
    qr.add_data(qr_code)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img_path = os.path.join(QR_DIR, f"{qr_code}.png")
    img.save(img_path)
    return img_path


def generate_card_image(pid, qr_code, nama, event_name):
    qr_path = os.path.join(QR_DIR, f"{qr_code}.png")
    if not os.path.exists(qr_path):
        generate_qr_image(pid, qr_code, nama)

    qr_img = PILImage.open(qr_path)
    card_w, card_h = 400, 520
    card = PILImage.new("RGB", (card_w, card_h), "white")

    qr_size = 220
    qr_resized = qr_img.resize((qr_size, qr_size))
    qr_x = (card_w - qr_size) // 2
    qr_y = 40
    card.paste(qr_resized, (qr_x, qr_y))

    draw = ImageDraw.Draw(card)

    try:
        font_large = ImageFont.truetype("arial.ttf", 22)
        font_medium = ImageFont.truetype("arial.ttf", 16)
        font_small = ImageFont.truetype("arial.ttf", 13)
    except OSError:
        font_large = ImageFont.load_default()
        font_medium = font_large
        font_small = font_large

    y_text = qr_y + qr_size + 20

    bbox = draw.textbbox((0, 0), event_name, font=font_large)
    tw = bbox[2] - bbox[0]
    draw.text(((card_w - tw) // 2, y_text), event_name, fill="black", font=font_large)
    y_text += 35

    bbox = draw.textbbox((0, 0), pid, font=font_small)
    tw = bbox[2] - bbox[0]
    draw.text(((card_w - tw) // 2, y_text), pid, fill="gray", font=font_small)
    y_text += 25

    bbox = draw.textbbox((0, 0), nama, font=font_medium)
    tw = bbox[2] - bbox[0]
    draw.text(((card_w - tw) // 2, y_text), nama, fill="black", font=font_medium)

    return card


def generate_qr_images_for_all():
    peserta = get_peserta_list()
    for p in peserta:
        qr_path = os.path.join(QR_DIR, f"{p['qrcode']}.png")
        if not os.path.exists(qr_path):
            generate_qr_image(p["id"], p["qrcode"], p["nama"])


def send_registration_email(to_email, pid, nama, qr_code, event_name):
    config = get_config()
    smtp_server = config.get("email_server", "smtp.gmail.com")
    smtp_port_raw = config.get("email_port", "587")
    smtp_port = int(smtp_port_raw) if str(smtp_port_raw).isdigit() else 587
    sender_email = config.get("email_sender", "")
    sender_password = config.get("email_password", "")

    if not sender_email or not sender_password:
        return False, "Email belum dikonfigurasi"

    qr_path = os.path.join(QR_DIR, f"{qr_code}.png")
    if not os.path.exists(qr_path):
        generate_qr_image(pid, qr_code, nama)

    msg = MIMEMultipart()
    msg["From"] = f"{event_name} <{sender_email}>"
    msg["To"] = to_email
    msg["Subject"] = f"Pendaftaran Berhasil - {event_name}"

    html = f"""
    <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto; padding: 20px;">
        <div style="background: #007bff; color: white; text-align: center; padding: 20px; border-radius: 12px 12px 0 0;">
            <h2 style="margin: 0;">{event_name}</h2>
        </div>
        <div style="background: #ffffff; padding: 30px; border: 1px solid #e0e0e0;">
            <p style="font-size: 16px;">Halo <strong>{nama}</strong>,</p>
            <p style="font-size: 16px;">Kamu sudah berhasil terdaftar! Berikut adalah data kamu:</p>
            <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; margin: 20px 0;">
                <p style="margin: 0; font-size: 14px; color: #666;">ID Peserta</p>
                <p style="margin: 5px 0; font-size: 28px; font-weight: bold; color: #198754;">{pid}</p>
            </div>
            <p style="font-size: 14px; color: #666; text-align: center;">Kode QR kamu terlampir di email ini.</p>
            <hr style="border: none; border-top: 1px solid #eee;">
            <p style="font-size: 13px; color: #999; text-align: center;">
                Simpan email ini dan tunjukkan ID Peserta saat check-in di hari H.
            </p>
        </div>
    </div>
    """

    msg.attach(MIMEText(html, "html"))

    if os.path.exists(qr_path):
        with open(qr_path, "rb") as f:
            img = MIMEImage(f.read())
            img.add_header("Content-Disposition", "inline", filename=f"QR_{qr_code}.png")
            msg.attach(img)

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, to_email, msg.as_string())
        return True, "Berhasil"
    except Exception as e:
        return False, str(e)


# --- Sesi Absensi ---

def get_sesi_list():
    sesi = []
    if os.path.exists(SESI_FILE):
        wb = load_workbook(SESI_FILE)
        ws = wb["Sesi"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                sesi.append({
                    "id": row[0],
                    "nama": row[1],
                    "aktif": row[2],
                })
        wb.close()
    return sesi


def get_sesi_by_id(sesi_id):
    for s in get_sesi_list():
        if s["id"] == sesi_id:
            return s
    return None


def get_active_sesi():
    for s in get_sesi_list():
        if s["aktif"]:
            return s
    return None


def get_next_sesi_id():
    sesi = get_sesi_list()
    if not sesi:
        return 1
    return max(s["id"] for s in sesi) + 1


def add_sesi(nama):
    sid = get_next_sesi_id()
    wb = load_workbook(SESI_FILE)
    ws = wb["Sesi"]
    ws.append([sid, nama, False])
    wb.save(SESI_FILE)
    return sid


def delete_sesi(sesi_id):
    wb = load_workbook(SESI_FILE)
    ws = wb["Sesi"]
    rows_to_delete = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == sesi_id:
            rows_to_delete.append(idx)
    for idx in reversed(rows_to_delete):
        ws.delete_rows(idx)
    wb.save(SESI_FILE)

    if os.path.exists(ABSENSI_FILE):
        wb2 = load_workbook(ABSENSI_FILE)
        ws2 = wb2["Absensi"]
        rows_to_delete2 = []
        for idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            if row[3] == sesi_id:
                rows_to_delete2.append(idx)
        for idx in reversed(rows_to_delete2):
            ws2.delete_rows(idx)
        wb2.save(ABSENSI_FILE)


def activate_sesi(sesi_id):
    wb = load_workbook(SESI_FILE)
    ws = wb["Sesi"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sesi_id:
            row[2].value = True
        else:
            row[2].value = False
    wb.save(SESI_FILE)


def deactivate_all_sesi():
    wb = load_workbook(SESI_FILE)
    ws = wb["Sesi"]
    for row in ws.iter_rows(min_row=2):
        row[2].value = False
    wb.save(SESI_FILE)


# --- Absensi ---

def get_absensi_list(sesi_id=None):
    absensi = []
    if os.path.exists(ABSENSI_FILE):
        wb = load_workbook(ABSENSI_FILE)
        ws = wb["Absensi"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                if sesi_id is not None and row[3] != sesi_id:
                    continue
                absensi.append({
                    "id_peserta": row[0],
                    "nama_peserta": row[1],
                    "divisi": row[2],
                    "id_sesi": row[3],
                    "waktu": row[4],
                    "tanggal": row[5],
                })
        wb.close()
    return absensi


def is_already_checked_in(peserta_id, sesi_id):
    absensi = get_absensi_list(sesi_id)
    for a in absensi:
        if a["id_peserta"] == peserta_id:
            return True
    return False


def log_audit(nama_peserta, id_peserta, aksi, sesi, status, detail=""):
    now = datetime.now()
    waktu = now.strftime("%Y-%m-%d %H:%M:%S")
    wb = load_workbook(AUDIT_LOG_FILE)
    ws = wb["AuditLog"]
    ws.append([waktu, id_peserta, nama_peserta, aksi, sesi, status, detail])
    wb.save(AUDIT_LOG_FILE)
    wb.close()


def checkin_peserta_multi(qr_code):
    today = date.today().isoformat()
    config = get_config()
    now = datetime.now().strftime("%H:%M")

    active = get_active_sesi()
    if not active:
        log_audit("-", "-", "Scan", "-", "Error", "Tidak ada sesi aktif")
        return {"status": "error", "message": "Tidak ada sesi absensi yang aktif!"}

    checkin_start = config.get("checkin_start", "08:00")
    checkin_end = config.get("checkin_end", "10:00")

    peserta = get_peserta_list()
    target = None
    for p in peserta:
        if p["qrcode"] == qr_code:
            target = p
            break

    if not target:
        log_audit("-", "-", "Scan", active["nama"], "Error", f"QR tidak valid: {qr_code[:8]}...")
        return {"status": "error", "message": "QR Code tidak valid!"}

    if is_already_checked_in(target["id"], active["id"]):
        log_audit(target["nama"], target["id"], "Scan", active["nama"], "Warning", "Sudah check-in")
        return {"status": "warning", "message": f"{target['nama']} sudah check-in di sesi {active['nama']}."}

    late = now > checkin_end

    wb = load_workbook(ABSENSI_FILE)
    ws = wb["Absensi"]
    ws.append([
        target["id"],
        target["nama"],
        target["divisi"],
        active["id"],
        datetime.now().strftime("%H:%M:%S"),
        today
    ])
    wb.save(ABSENSI_FILE)

    msg = f"{target['nama']} berhasil check-in ({active['nama']})!"
    if late:
        msg = f"{target['nama']} berhasil check-in ({active['nama']}, TERLAMBAT)!"
        log_audit(target["nama"], target["id"], "Check-in", active["nama"], "Success", "Terlambat")
    else:
        log_audit(target["nama"], target["id"], "Check-in", active["nama"], "Success", "")
    return {"status": "success", "message": msg, "late": late, "sesi": active["nama"]}


def get_stats_for_sesi(sesi_id=None):
    peserta = get_peserta_list()
    total = len(peserta)
    if sesi_id is not None:
        absensi = get_absensi_list(sesi_id)
        hadir = len(absensi)
    else:
        hadir = sum(1 for p in peserta if p["hadir"])
    belum = total - hadir
    persentase = round((hadir / total * 100), 1) if total > 0 else 0
    return {"total": total, "hadir": hadir, "belum": belum, "persentase": persentase}


# --- Routes ---

@app.route("/")
@login_required
def index():
    config = get_config()
    return render_template("index.html", config=config)


@app.route("/register", methods=["GET", "POST"])
def register():
    config = get_config()
    if config.get("registration_open", "true") != "true":
        return render_template("register_closed.html", config=config)
    if request.method == "POST":
        nama = request.form.get("nama", "").strip()
        email = request.form.get("email", "").strip()
        nohp = request.form.get("nohp", "").strip()
        divisi = request.form.get("divisi", "").strip()
        if not nama:
            flash("Nama wajib diisi!", "danger")
            return redirect(url_for("register"))
        if not email:
            flash("Email wajib diisi untuk menerima ID dan QR Code!", "danger")
            return redirect(url_for("register"))
        pid, qr_code = add_peserta(nama, email, nohp, divisi)
        event_name = config.get("event_name", "Absensi Event")
        email_ok, email_msg = send_registration_email(email, pid, nama, qr_code, event_name)
        return render_template("register_success.html", pid=pid, nama=nama,
                               email=email, email_ok=email_ok, email_msg=email_msg, config=config)
    return render_template("register.html", config=config)


@app.route("/scan", methods=["GET", "POST"])
@login_required
def scan():
    if request.method == "POST":
        qr_code = request.form.get("qr_code", "").strip()
        if qr_code:
            result = checkin_peserta_multi(qr_code)
            return jsonify(result)
        return jsonify({"status": "error", "message": "Kode QR kosong!"})
    config = get_config()
    active_sesi = get_active_sesi()
    return render_template("scan.html", config=config, active_sesi=active_sesi)


@app.route("/api/checkin", methods=["POST"])
@login_required
def api_checkin():
    data = request.get_json()
    qr_code = data.get("qr_code", "").strip() if data else ""
    if not qr_code:
        return jsonify({"status": "error", "message": "Kode QR kosong!"})
    result = checkin_peserta_multi(qr_code)
    return jsonify(result)


@app.route("/api/active-session")
@login_required
def api_active_session():
    active = get_active_sesi()
    if active:
        return jsonify({"id": active["id"], "nama": active["nama"]})
    return jsonify(None)


@app.route("/api/stats")
@login_required
def api_stats():
    active = get_active_sesi()
    if active:
        return jsonify(get_stats_for_sesi(active["id"]))
    return jsonify(get_stats_for_sesi())


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        if username == ADMIN_USER and password == ADMIN_PASS:
            session["admin_logged_in"] = True
            return redirect(url_for("admin_dashboard"))
        flash("Username atau password salah!", "danger")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login"))


@app.route("/admin/dashboard")
@login_required
def admin_dashboard():
    config = get_config()
    sesi_list = get_sesi_list()
    active_sesi = get_active_sesi()
    filter_sesi = request.args.get("sesi", type=int)

    if filter_sesi:
        stats = get_stats_for_sesi(filter_sesi)
        absensi = get_absensi_list(filter_sesi)
        filter_nama = get_sesi_by_id(filter_sesi)
    elif active_sesi:
        stats = get_stats_for_sesi(active_sesi["id"])
        absensi = get_absensi_list(active_sesi["id"])
        filter_nama = active_sesi
    else:
        stats = get_stats_for_sesi()
        absensi = []
        filter_nama = None

    return render_template("admin/dashboard.html",
                           stats=stats, absensi=absensi,
                           sesi_list=sesi_list, active_sesi=active_sesi,
                           filter_nama=filter_nama, config=config)


@app.route("/admin/peserta")
@login_required
def admin_peserta():
    peserta = get_peserta_list()
    return render_template("admin/peserta.html", peserta=peserta)


@app.route("/admin/peserta/tambah", methods=["GET", "POST"])
@login_required
def admin_tambah_peserta():
    if request.method == "POST":
        nama = request.form.get("nama", "").strip()
        email = request.form.get("email", "").strip()
        nohp = request.form.get("nohp", "").strip()
        divisi = request.form.get("divisi", "").strip()
        if nama:
            pid, qr_code = add_peserta(nama, email, nohp, divisi)
            flash(f"Peserta {nama} berhasil ditambahkan! ID: {pid}", "success")
            return redirect(url_for("admin_peserta"))
        flash("Nama wajib diisi!", "danger")
    return render_template("admin/tambah_peserta.html")


@app.route("/admin/peserta/hapus/<pid>")
@login_required
def admin_hapus_peserta(pid):
    delete_peserta(pid)
    flash("Peserta berhasil dihapus!", "success")
    return redirect(url_for("admin_peserta"))


@app.route("/admin/peserta/import", methods=["GET", "POST"])
@login_required
def admin_import_peserta():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("Pilih file terlebih dahulu!", "danger")
            return redirect(url_for("admin_import_peserta"))

        filename = file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
            flash("Format file harus .xlsx atau .csv!", "danger")
            return redirect(url_for("admin_import_peserta"))

        try:
            if filename.endswith(".csv"):
                import csv
                import codecs
                stream = codecs.iterdecode(file.stream, "utf-8-sig")
                reader = csv.reader(stream)
                rows = list(reader)
            else:
                wb = load_workbook(file)
                ws = wb.active
                rows = []
                for row in ws.iter_rows(min_row=1, values_only=True):
                    rows.append(list(row))
                wb.close()

            if not rows:
                flash("File kosong!", "danger")
                return redirect(url_for("admin_import_peserta"))

            header = [str(h).strip().lower() if h else "" for h in rows[0]]
            name_idx = None
            email_idx = None
            nohp_idx = None
            divisi_idx = None
            for i, h in enumerate(header):
                if h in ("nama", "name"):
                    name_idx = i
                elif h in ("email",):
                    email_idx = i
                elif h in ("nohp", "no hp", "no. hp", "phone", "telepon", "whatsapp", "wa"):
                    nohp_idx = i
                elif h in ("divisi", "kelompok", "group", "department"):
                    divisi_idx = i

            if name_idx is None:
                flash("File harus memiliki kolom 'Nama'!", "danger")
                return redirect(url_for("admin_import_peserta"))

            success_count = 0
            fail_count = 0
            for row in rows[1:]:
                if not row or not any(row):
                    continue
                nama = str(row[name_idx]).strip() if name_idx is not None and name_idx < len(row) and row[name_idx] else ""
                if not nama:
                    fail_count += 1
                    continue
                email = str(row[email_idx]).strip() if email_idx is not None and email_idx < len(row) and row[email_idx] else ""
                nohp = str(row[nohp_idx]).strip() if nohp_idx is not None and nohp_idx < len(row) and row[nohp_idx] else ""
                divisi = str(row[divisi_idx]).strip() if divisi_idx is not None and divisi_idx < len(row) and row[divisi_idx] else ""
                add_peserta(nama, email, nohp, divisi)
                success_count += 1

            flash(f"Import selesai! {success_count} peserta berhasil ditambahkan, {fail_count} gagal.", "success")
        except Exception as e:
            flash(f"Gagal membaca file: {str(e)}", "danger")

        return redirect(url_for("admin_peserta"))

    return render_template("admin/import_peserta.html")


@app.route("/admin/sesi")
@login_required
def admin_sesi():
    sesi_list = get_sesi_list()
    active_sesi = get_active_sesi()
    return render_template("admin/sesi.html", sesi_list=sesi_list, active_sesi=active_sesi)


@app.route("/admin/sesi/tambah", methods=["POST"])
@login_required
def admin_tambah_sesi():
    nama = request.form.get("nama", "").strip()
    if nama:
        add_sesi(nama)
        flash(f"Sesi \"{nama}\" berhasil ditambahkan!", "success")
    else:
        flash("Nama sesi wajib diisi!", "danger")
    return redirect(url_for("admin_sesi"))


@app.route("/admin/sesi/hapus/<int:sid>")
@login_required
def admin_hapus_sesi(sid):
    delete_sesi(sid)
    flash("Sesi berhasil dihapus!", "success")
    return redirect(url_for("admin_sesi"))


@app.route("/admin/sesi/activate/<int:sid>")
@login_required
def admin_activate_sesi(sid):
    activate_sesi(sid)
    sesi = get_sesi_by_id(sid)
    flash(f"Sesi \"{sesi['nama']}\" sekarang aktif!", "success")
    return redirect(url_for("admin_sesi"))


@app.route("/admin/cetak")
@login_required
def admin_cetak():
    peserta = get_peserta_list()
    config = get_config()
    return render_template("admin/cetak_qr.html", peserta=peserta, config=config)


@app.route("/api/qr-image/<qr_code>")
def qr_image(qr_code):
    qr_path = os.path.join(QR_DIR, f"{qr_code}.png")
    if os.path.exists(qr_path):
        return send_from_directory(QR_DIR, f"{qr_code}.png", mimetype="image/png")
    return "Not found", 404


@app.route("/api/card-image/<pid>")
def card_image(pid):
    peserta = get_peserta_list()
    config = get_config()
    event_name = config.get("event_name", "Event")
    target = None
    for p in peserta:
        if p["id"] == pid:
            target = p
            break
    if not target:
        return "Not found", 404

    card = generate_card_image(target["id"], target["qrcode"], target["nama"], event_name)
    buf = io.BytesIO()
    card.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png", download_name=f"card_{target['id']}.png")


@app.route("/api/download-qr/<qr_code>")
def download_qr(qr_code):
    qr_path = os.path.join(QR_DIR, f"{qr_code}.png")
    if not os.path.exists(qr_path):
        generate_qr_images_for_all()
    if os.path.exists(qr_path):
        return send_from_directory(QR_DIR, f"{qr_code}.png", as_attachment=True,
                                   download_name=f"QR_{qr_code}.png")
    return "Not found", 404


@app.route("/api/download-card/<pid>")
def download_card(pid):
    peserta = get_peserta_list()
    config = get_config()
    event_name = config.get("event_name", "Event")
    target = None
    for p in peserta:
        if p["id"] == pid:
            target = p
            break
    if not target:
        return "Not found", 404

    card = generate_card_image(target["id"], target["qrcode"], target["nama"], event_name)
    buf = io.BytesIO()
    card.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png", as_attachment=True,
                     download_name=f"Kartu_{target['id']}_{target['nama']}.png")


@app.route("/api/download-all-cards-pdf")
@login_required
def download_all_cards_pdf():
    peserta = get_peserta_list()
    config = get_config()
    event_name = config.get("event_name", "Event")

    if not peserta:
        return "Tidak ada peserta", 404

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    page_w, page_h = A4
    card_w = 60 * mm
    card_h = 80 * mm
    cols = 3
    rows = 3
    margin_x = 15 * mm
    margin_y = 15 * mm
    gap_x = (page_w - 2 * margin_x - cols * card_w) / (cols - 1)
    gap_y = (page_h - 2 * margin_y - rows * card_h) / (rows - 1)

    per_page = cols * rows
    page_num = 0

    for i, p in enumerate(peserta):
        pos_on_page = i % per_page
        if pos_on_page == 0 and i > 0:
            c.showPage()
            page_num += 1

        col = pos_on_page % cols
        row = pos_on_page // cols
        x = margin_x + col * (card_w + gap_x)
        y = page_h - margin_y - card_h - row * (card_h + gap_y)

        c.setStrokeColor("#cccccc")
        c.rect(x, y, card_w, card_h)

        qr_path = os.path.join(QR_DIR, f"{p['qrcode']}.png")
        if not os.path.exists(qr_path):
            generate_qr_image(p["id"], p["qrcode"], p["nama"])

        if os.path.exists(qr_path):
            qr_size = 40 * mm
            qr_x = x + (card_w - qr_size) / 2
            qr_y = y + card_h - qr_size - 8 * mm
            c.drawImage(ImageReader(qr_path), qr_x, qr_y, qr_size, qr_size)

        c.setFont("Helvetica-Bold", 8)
        text_x = x + card_w / 2
        c.drawCentredString(text_x, y + 22 * mm, p["id"])
        c.setFont("Helvetica", 7)
        c.drawCentredString(text_x, y + 15 * mm, p["nama"])

        c.setFont("Helvetica", 5)
        c.drawCentredString(text_x, y + 5 * mm, event_name)

    c.save()
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name="Semua_Kartu_QR.pdf")


@app.route("/admin/settings", methods=["GET", "POST"])
@login_required
def admin_settings():
    config = get_config()
    if request.method == "POST":
        config["event_name"] = request.form.get("event_name", config.get("event_name", ""))
        config["event_date"] = request.form.get("event_date", config.get("event_date", ""))
        config["event_location"] = request.form.get("event_location", config.get("event_location", ""))
        config["checkin_start"] = request.form.get("checkin_start", config.get("checkin_start", "08:00"))
        config["checkin_end"] = request.form.get("checkin_end", config.get("checkin_end", "10:00"))
        config["checkout_start"] = request.form.get("checkout_start", config.get("checkout_start", "16:00"))
        config["checkout_end"] = request.form.get("checkout_end", config.get("checkout_end", "18:00"))
        config["notif_sound"] = "true" if request.form.get("notif_sound") else "false"
        config["registration_open"] = "true" if request.form.get("registration_open") else "false"
        config["email_server"] = request.form.get("email_server", config.get("email_server", "smtp.gmail.com"))
        email_port = request.form.get("email_port", "587").strip()
        config["email_port"] = email_port if email_port.isdigit() else "587"
        config["email_sender"] = request.form.get("email_sender", config.get("email_sender", ""))
        config["email_password"] = request.form.get("email_password", config.get("email_password", ""))
        save_config(config)
        flash("Pengaturan berhasil disimpan!", "success")
        return redirect(url_for("admin_settings"))
    return render_template("admin/settings.html", config=config)


@app.route("/admin/export-excel")
@login_required
def admin_export_excel():
    sesi_list = get_sesi_list()
    return render_template("admin/export.html", sesi_list=sesi_list)


@app.route("/admin/export-excel/download/<int:sid>")
@login_required
def admin_export_excel_download(sid):
    sesi = get_sesi_by_id(sid)
    if not sesi:
        flash("Sesi tidak ditemukan!", "danger")
        return redirect(url_for("admin_export_excel"))

    config = get_config()
    event_name = config.get("event_name", "Event")
    absensi = get_absensi_list(sid)

    wb = Workbook()

    ws = wb.active
    ws.title = sesi["nama"][:31]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["No", "ID Peserta", "Nama", "Divisi", "Waktu Check-In", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    checkin_end = config.get("checkin_end", "10:00")
    for i, a in enumerate(absensi, 1):
        is_late = a["waktu"][:5] > checkin_end if a["waktu"] else False
        status = "Terlambat" if is_late else "Hadir"
        row_data = [i, a["id_peserta"], a["nama_peserta"], a["divisi"] or "-", a["waktu"], status]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if status == "Terlambat" and col == 6:
                cell.font = Font(color="FF0000", bold=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14

    ws2 = wb.create_sheet("Ringkasan")
    ws2.cell(row=1, column=1, value="Rekapitulasi Kehadiran").font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")

    ws2.cell(row=3, column=1, value="Nama Event").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=event_name)
    ws2.cell(row=4, column=1, value="Sesi Absensi").font = Font(bold=True)
    ws2.cell(row=4, column=2, value=sesi["nama"])
    ws2.cell(row=5, column=1, value="Tanggal Export").font = Font(bold=True)
    ws2.cell(row=5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    stats = get_stats_for_sesi(sid)
    ws2.cell(row=7, column=1, value="Total Peserta").font = Font(bold=True)
    ws2.cell(row=7, column=2, value=stats["total"])
    ws2.cell(row=8, column=1, value="Hadir").font = Font(bold=True)
    ws2.cell(row=8, column=2, value=stats["hadir"])
    ws2.cell(row=9, column=1, value="Belum Hadir").font = Font(bold=True)
    ws2.cell(row=9, column=2, value=stats["belum"])
    ws2.cell(row=10, column=1, value="Persentase").font = Font(bold=True)
    ws2.cell(row=10, column=2, value=f"{stats['persentase']}%")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Hasil_{event_name}_{sesi['nama']}.xlsx".replace(" ", "_")
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/admin/audit-log")
@login_required
def admin_audit_log():
    logs = []
    if os.path.exists(AUDIT_LOG_FILE):
        wb = load_workbook(AUDIT_LOG_FILE)
        ws = wb["AuditLog"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                logs.append({
                    "waktu": row[0],
                    "id_peserta": row[1],
                    "nama_peserta": row[2],
                    "aksi": row[3],
                    "sesi": row[4],
                    "status": row[5],
                    "detail": row[6],
                })
        wb.close()
    logs.reverse()
    return render_template("admin/audit_log.html", logs=logs)


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
